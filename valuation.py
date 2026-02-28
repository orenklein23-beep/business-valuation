import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Valuation Model"

# Colors
dark_blue = "1a1a2e"
medium_blue = "16213e"
light_blue = "0f3460"
green = "00b300"
light_gray = "f2f2f2"
white = "ffffff"
yellow = "ffd700"

def style_cell(cell, bold=False, size=11, color="000000", bg=None, align="left", wrap=False):
    cell.font = Font(bold=bold, size=size, color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def border_cell(cell):
    thin = Side(style="thin", color="cccccc")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20

# Row heights
for i in range(1, 80):
    ws.row_dimensions[i].height = 20

# ── TITLE ──
ws.merge_cells('A1:E1')
ws['A1'] = "MAIN STREET OUTFITTERS — BUSINESS VALUATION"
style_cell(ws['A1'], bold=True, size=16, color=white, bg=dark_blue, align="center")
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:E2')
ws['A2'] = "Local Retail Clothing Boutique | New York, NY | FY 2024"
style_cell(ws['A2'], size=11, color=white, bg=medium_blue, align="center")

# ── SECTION: BUSINESS ASSUMPTIONS ──
ws.merge_cells('A4:E4')
ws['A4'] = "BUSINESS ASSUMPTIONS"
style_cell(ws['A4'], bold=True, size=12, color=white, bg=light_blue, align="center")

headers = ["Metric", "Value", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    style_cell(c, bold=True, color=white, bg=light_blue, align="center")

assumptions = [
    ("Annual Revenue", 1200000, "Based on avg boutique with 2,000 sq ft"),
    ("Revenue Growth Rate", "5%", "Conservative for stable retail"),
    ("Gross Margin", "55%", "Typical for clothing retail"),
    ("Operating Expenses", 480000, "Rent, staff, utilities, marketing"),
    ("EBITDA", 180000, "Revenue x Gross Margin - OpEx"),
    ("Net Profit Margin", "15%", "EBITDA / Revenue"),
    ("Net Profit", 180000, "Annual after-tax earnings estimate"),
    ("Discount Rate (WACC)", "12%", "Risk-adjusted for small business"),
    ("Terminal Growth Rate", "2%", "Long-term stable growth"),
]

for i, (metric, value, note) in enumerate(assumptions, 6):
    ws.cell(row=i, column=1, value=metric)
    style_cell(ws.cell(row=i, column=1), bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=1))

    ws.cell(row=i, column=2, value=value)
    style_cell(ws.cell(row=i, column=2), bold=True, align="center", bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=2))

    ws.cell(row=i, column=3, value=note)
    style_cell(ws.cell(row=i, column=3), size=10, color="666666", bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=3))

# ── SECTION: METHOD 1 - REVENUE MULTIPLE ──
ws.merge_cells('A17:E17')
ws['A17'] = "METHOD 1: REVENUE MULTIPLE"
style_cell(ws['A17'], bold=True, size=12, color=white, bg=light_blue, align="center")

rev_data = [
    ("Annual Revenue", "$1,200,000"),
    ("Industry Revenue Multiple", "0.5x — 1.0x"),
    ("Low Estimate (0.5x)", "$600,000"),
    ("High Estimate (1.0x)", "$1,200,000"),
    ("Mid Estimate", "$900,000"),
]
for i, (label, value) in enumerate(rev_data, 18):
    ws.cell(row=i, column=1, value=label)
    style_cell(ws.cell(row=i, column=1), bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=1))
    ws.cell(row=i, column=2, value=value)
    style_cell(ws.cell(row=i, column=2), bold=True, align="center", bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=2))

# ── SECTION: METHOD 2 - EARNINGS MULTIPLE ──
ws.merge_cells('A25:E25')
ws['A25'] = "METHOD 2: EARNINGS MULTIPLE (EV/EBITDA)"
style_cell(ws['A25'], bold=True, size=12, color=white, bg=light_blue, align="center")

earn_data = [
    ("EBITDA", "$180,000"),
    ("Industry EBITDA Multiple", "3x — 5x"),
    ("Low Estimate (3x)", "$540,000"),
    ("High Estimate (5x)", "$900,000"),
    ("Mid Estimate", "$720,000"),
]
for i, (label, value) in enumerate(earn_data, 26):
    ws.cell(row=i, column=1, value=label)
    style_cell(ws.cell(row=i, column=1), bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=1))
    ws.cell(row=i, column=2, value=value)
    style_cell(ws.cell(row=i, column=2), bold=True, align="center", bg=light_gray if i % 2 == 0 else white)
    border_cell(ws.cell(row=i, column=2))

# ── SECTION: METHOD 3 - DCF ──
ws.merge_cells('A33:E33')
ws['A33'] = "METHOD 3: DISCOUNTED CASH FLOW (DCF)"
style_cell(ws['A33'], bold=True, size=12, color=white, bg=light_blue, align="center")

dcf_headers = ["Year", "Cash Flow", "Growth", "Discount Factor", "PV of Cash Flow"]
for col, h in enumerate(dcf_headers, 1):
    c = ws.cell(row=34, column=col, value=h)
    style_cell(c, bold=True, color=white, bg=light_blue, align="center")

cf = 180000
wacc = 0.12
growth = 0.05
total_pv = 0
for year in range(1, 6):
    cf = cf * (1 + growth)
    discount = (1 + wacc) ** year
    pv = cf / discount
    total_pv += pv
    row = 34 + year
    ws.cell(row=row, column=1, value=f"Year {year}")
    ws.cell(row=row, column=2, value=round(cf, 0))
    ws.cell(row=row, column=3, value=f"{growth*100:.0f}%")
    ws.cell(row=row, column=4, value=round(discount, 3))
    ws.cell(row=row, column=5, value=round(pv, 0))
    for col in range(1, 6):
        style_cell(ws.cell(row=row, column=col), bg=light_gray if year % 2 == 0 else white, align="center")
        border_cell(ws.cell(row=row, column=col))

# Terminal value
terminal_cf = cf * (1 + 0.02)
terminal_value = terminal_cf / (wacc - 0.02)
pv_terminal = terminal_value / (1 + wacc) ** 5
total_value = total_pv + pv_terminal

ws.cell(row=40, column=1, value="Terminal Value (PV)")
ws.cell(row=40, column=5, value=round(pv_terminal, 0))
ws.cell(row=41, column=1, value="Sum of PV of Cash Flows")
ws.cell(row=41, column=5, value=round(total_pv, 0))
ws.cell(row=42, column=1, value="DCF Valuation")
ws.cell(row=42, column=5, value=round(total_value, 0))

for row in [40, 41, 42]:
    for col in [1, 5]:
        style_cell(ws.cell(row=row, column=col), bold=True, bg=light_gray, align="center")
        border_cell(ws.cell(row=row, column=col))

# ── SECTION: FINAL VALUATION SUMMARY ──
ws.merge_cells('A45:E45')
ws['A45'] = "FINAL VALUATION SUMMARY"
style_cell(ws['A45'], bold=True, size=12, color=white, bg=dark_blue, align="center")

summary_headers = ["Method", "Low", "Mid", "High"]
for col, h in enumerate(summary_headers, 1):
    c = ws.cell(row=46, column=col, value=h)
    style_cell(c, bold=True, color=white, bg=light_blue, align="center")

summary = [
    ("Revenue Multiple", "$600,000", "$900,000", "$1,200,000"),
    ("Earnings Multiple", "$540,000", "$720,000", "$900,000"),
    (f"DCF Analysis", f"${round(total_value*0.9):,}", f"${round(total_value):,}", f"${round(total_value*1.1):,}"),
]
for i, (method, low, mid, high) in enumerate(summary, 47):
    for col, val in enumerate([method, low, mid, high], 1):
        c = ws.cell(row=i, column=col, value=val)
        style_cell(c, bg=light_gray if i % 2 == 0 else white, align="center")
        border_cell(c)

ws.merge_cells('A51:E51')
ws['A51'] = f"ESTIMATED VALUATION RANGE: $600,000 — $1,200,000"
style_cell(ws['A51'], bold=True, size=13, color=white, bg="00b300", align="center")
ws.row_dimensions[51].height = 30

wb.save("MainStreet_Valuation.xlsx")
print("Done! Saved as MainStreet_Valuation.xlsx")